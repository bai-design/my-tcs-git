import openpyxl

class Write_excel_xlsx_append(object):
    def __init__(self, path, value, sheet_name):
        self.path = path
        self.value = value
        self.sheet_name = sheet_name

    def write_excel_xlsx_append(self):
        """向xls文件内写入（追加）内容"""

        # 写入内容格式[[], [],[]..]
        # 写入内容行数
        index = len(self.value)
        # 读模式打开文件
        work_book = openpyxl.load_workbook(self.path)
        # 定位sheet页
        work_sheet = work_book[self.sheet_name]
        # sheet页数据行数
        rows_old = work_sheet.max_row
        rows = lambda x: x-1 if x else 0
        row_old_real = rows(rows_old)
        print('原sheet页数据行数为: %s' % (row_old_real))
        # 追加内容
        for i in range(1, index+1):
            for j in range(1, len(self.value[i-1])+1):
                work_sheet.cell(rows_old+i-1, j).value = self.value[i-1][j-1]
        print('写入完成，%s页数据行数为: %s' % (self.sheet_name, rows_old + index))
        # 保存文件
        work_book.save(self.path)
        print('新数据保存成功')

